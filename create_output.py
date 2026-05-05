#!/usr/bin/env python3
"""
Double-click helper for Pella order generation.

Place this file in the same folder as:
  - purchase_order.csv
  - work_order.xlsx

Then double-click it, or run:
  python create_output.py

The script creates output.xlsx in the same folder.
"""

from __future__ import annotations

import csv
import re
import sys
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path


PO_FILENAME = "purchase_order.csv"
WORK_ORDER_FILENAME = "work_order.xlsx"
OUTPUT_FILENAME = "output.xlsx"

FINISH_CODES = {
    "unfinished": "NONE",
    "primed": "PRIME",
    "paint pella white": "PWHITE",
    "pella white": "PWHITE",
    "paint vinyl white": "VWHITE",
    "vinyl white": "VWHITE",
    "paint pella black": "PBLACK",
    "pella black": "PBLACK",
    "paint linen white": "LINENW",
    "linen white": "LINENW",
    "paint bright white": "BRIGHTW",
    "bright white": "BRIGHTW",
}

INPUT_HEADERS = [
    "Type",
    "PO Line",
    "Pella Xref",
    "Wall Depth(JambDepth)",
    "Window Depth",
    "Jamb Depth",
    "Jamb Width",
    "Jamb height",
    "Quantity",
    "Jamb Style",
    "Lumber Type",
    "Finish",
    "Calculated Jamb height",
    "Calculated Footage",
]

OUTPUT_HEADERS = [
    "Date",
    "Schedule",
    "Batch(Box =)",
    "Bin",
    "Part",
    "Loc",
    "W",
    "Fin",
    "Qty",
    "L",
    "Info",
    "Order NumLer",
    "Assembled",
    "Debug",
    "Part No",
    "Customer",
    "Shp Addr City",
    "Shp Addr State",
    "PONumber",
    "Customer Ref",
    "Comment",
    "Shp Addr Address1",
    "Shp Addr Zip Code",
    "orderline",
    "Req Date",
    "SUMP",
    "Shim SUMP",
]

LINEAL_HEADERS = [
    "Date",
    "Schedule",
    "Batch(Box #)",
    "Bin",
    "Part",
    "Loc",
    "W",
    "Fin",
    "Qty",
    "L",
    "Info",
    "Part No",
    "Customer",
    "Shp Addr City",
    "Shp Addr State",
    "PONumber",
    "Customer Ref",
    "Comment",
    "Shp Addr Add1",
    "Shp Addr Zip",
    "Order Number",
    "Line Item",
    "Req Date",
    "SUMP",
]

LUMBER_TYPES = ("Poplar", "Pine", "Red Oak", "Maple")
JAMB_THICKNESS = 0.6875
XREF_RE = re.compile(r"(?P<order>[A-Z0-9]+)-(?P<suffix>\d{3,})")
MIXED_FRACTION_RE = re.compile(
    r"^\s*(?P<whole>\d+)?(?:[-\s]+)?(?:(?P<num>\d+)\s*/\s*(?P<den>\d+))?\s*$"
)


@dataclass
class WarningMessage:
    code: str
    message: str
    source: str | None = None


@dataclass
class PurchaseOrderLine:
    source_row: int
    po_line_number: str
    pella_xref: str
    xref_suffix: int
    description: str
    requested_date: date | None
    quantity: int | None
    pella_po_number: str


@dataclass
class WorkOrderItem:
    source_row: int
    item_number: int
    quantity: int | None
    rough_opening: str | None
    location: str | None
    comment: str | None
    description: str


@dataclass
class WorkOrder:
    path: Path
    order_number: str
    customer_name: str
    project_name: str
    items: dict[int, WorkOrderItem]


@dataclass
class ExtractedJambLine:
    po_line: PurchaseOrderLine
    work_item: WorkOrderItem | None
    type: str | None
    jamb_depth: float | None
    jamb_width: float | None
    jamb_height: float | None
    quantity: int
    piece_count: int | None
    jamb_style: str | None
    lumber_type: str | None
    finish_text: str | None
    finish_code: str | None
    calculated_jamb_height: float | None
    calculated_footage: float | None
    warnings: list[WarningMessage] = field(default_factory=list)


@dataclass
class JobModel:
    po_path: Path
    work_order_path: Path
    order_number: str
    pella_po_number: str
    customer_name: str
    ship_city: str = "Green Bay"
    ship_state: str = "WI"
    ship_address: str = "Pella"
    ship_zip: str = "54201"
    schedule: str | None = None
    date_code: str | None = None
    jamb_lines: list[ExtractedJambLine] = field(default_factory=list)
    warnings: list[WarningMessage] = field(default_factory=list)


def main() -> int:
    no_pause = "--no-pause" in sys.argv
    debug = "--debug" in sys.argv
    try:
        return run()
    except KeyboardInterrupt:
        print("\nCanceled.")
        return 130
    except ImportError as error:
        if "openpyxl" in str(error):
            print("ERROR: This script needs the Python package 'openpyxl'.")
            print("Install it with this command, then run the script again:")
            print("  python -m pip install openpyxl")
            return 1
        print(f"ERROR: Missing Python package: {error}")
        return 1
    except PermissionError as error:
        print(f"ERROR: {error}")
        print("If output.xlsx is open in Excel, close it and run this script again.")
        return 1
    except Exception as error:
        print(f"ERROR: {error}")
        if debug:
            traceback.print_exc()
        else:
            print("Run with --debug to show technical details.")
        return 1
    finally:
        if should_pause(no_pause):
            input("\nPress Enter to close this window...")


def run() -> int:
    folder = Path(__file__).resolve().parent
    po_path = find_file(folder, PO_FILENAME)
    work_order_path = find_file(folder, WORK_ORDER_FILENAME)
    output_path = folder / OUTPUT_FILENAME

    print("Pella output generator")
    print(f"Folder: {folder}")

    missing = []
    if po_path is None:
        missing.append(PO_FILENAME)
    if work_order_path is None:
        missing.append(WORK_ORDER_FILENAME)
    if missing:
        print("\nERROR: Missing required input file(s):")
        for filename in missing:
            print(f"  - {filename}")
        print("\nPlace the missing file(s) in the same folder as this script and try again.")
        return 1

    print(f"Purchase order: {po_path.name}")
    print(f"Work order:     {work_order_path.name}")
    print(f"Output:         {output_path.name}")

    job = build_job(po_path, work_order_path)
    written_path = write_workbook(job, output_path)
    warnings = dedupe_warnings(
        [*job.warnings, *(warning for line in job.jamb_lines for warning in line.warnings)]
    )

    print(f"\nDone. Created {written_path.name}.")
    print(f"Jamb lines: {len(job.jamb_lines)}")
    if warnings:
        print("\nWarnings to review:")
        for warning in warnings:
            print(f"  - [{warning.code}] {warning.message}")
    return 0


def should_pause(no_pause: bool) -> bool:
    if no_pause:
        return False
    if "--debug" in sys.argv:
        return False
    if len([arg for arg in sys.argv[1:] if not arg.startswith("--")]) > 0:
        return False
    return sys.stdin.isatty()


def find_file(folder: Path, expected_name: str) -> Path | None:
    exact = folder / expected_name
    if exact.is_file():
        return exact

    expected_lower = expected_name.lower()
    matches = [path for path in folder.iterdir() if path.is_file() and path.name.lower() == expected_lower]
    return matches[0] if matches else None


def build_job(po_path: str | Path, work_order_path: str | Path) -> JobModel:
    po_lines, po_warnings = parse_po_csv(po_path)
    work_order, wo_warnings = parse_work_order_xlsx(work_order_path)

    pella_po_number = po_lines[0].pella_po_number if po_lines else ""
    order_number = work_order.order_number or order_from_xref(po_lines[0].pella_xref if po_lines else "")
    job = JobModel(
        po_path=Path(po_path),
        work_order_path=Path(work_order_path),
        order_number=order_number,
        pella_po_number=pella_po_number,
        customer_name=work_order.project_name or work_order.customer_name,
        jamb_lines=[],
        warnings=[*po_warnings, *wo_warnings],
    )

    for po_line in po_lines:
        work_item = work_order.items.get(po_line.xref_suffix)
        if work_item is None:
            job.warnings.append(
                WarningMessage(
                    "match.no_work_item",
                    f"No work-order item matched PO xref {po_line.pella_xref}.",
                    f"PO row {po_line.source_row}",
                )
            )
        job.jamb_lines.extend(expand_po_line(po_line, work_item, work_order))

    return job


def parse_po_csv(path: str | Path) -> tuple[list[PurchaseOrderLine], list[WarningMessage]]:
    path = Path(path)
    warnings: list[WarningMessage] = []
    lines: list[PurchaseOrderLine] = []

    if path.suffix.lower() != ".csv":
        raise ValueError(f"Only CSV purchase orders are supported: {path.name}")

    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    for row_number, row in enumerate(rows, start=1):
        if not row:
            continue
        xref_index = find_xref_index(row)
        if xref_index is None:
            continue

        xref = row[xref_index].strip()
        match = XREF_RE.fullmatch(xref)
        if not match:
            continue

        try:
            suffix = int(match.group("suffix"))
        except ValueError:
            warnings.append(WarningMessage("po.bad_xref", f"Could not parse PO xref suffix from {xref}.", str(path)))
            continue

        po_number = value_after(row, "PO Number") or ""
        po_line_number = row[xref_index - 1].strip() if xref_index > 0 else ""
        description = row[xref_index + 1].strip() if xref_index + 1 < len(row) else ""
        requested_date = parse_short_date(row[xref_index + 2].strip() if xref_index + 2 < len(row) else "")
        quantity = parse_int(row[xref_index + 3].strip() if xref_index + 3 < len(row) else "")

        lines.append(
            PurchaseOrderLine(
                source_row=row_number,
                po_line_number=po_line_number,
                pella_xref=xref,
                xref_suffix=suffix,
                description=description,
                requested_date=requested_date,
                quantity=quantity,
                pella_po_number=po_number,
            )
        )

    if not lines:
        warnings.append(WarningMessage("po.no_lines", "No purchase-order lines were found.", str(path)))

    return lines, warnings


def parse_work_order_xlsx(path: str | Path) -> tuple[WorkOrder, list[WarningMessage]]:
    from openpyxl import load_workbook

    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only XLSX work orders are supported: {path.name}")

    warnings: list[WarningMessage] = []
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    order_number = clean(sheet["J1"].value)
    customer_name = clean(sheet["G1"].value)
    project_name = clean(sheet["H1"].value)

    if not order_number:
        warnings.append(WarningMessage("wo.no_order", "Could not find work-order number.", str(path)))

    items: dict[int, WorkOrderItem] = {}
    for row in range(2, sheet.max_row + 1):
        if clean(sheet.cell(row, 1).value) != "Item No.":
            continue
        item_number = parse_int(sheet.cell(row, 11).value)
        if item_number is None:
            continue
        items[item_number] = WorkOrderItem(
            source_row=row,
            item_number=item_number,
            quantity=parse_int(sheet.cell(row, 13).value),
            rough_opening=clean(sheet.cell(row, 8).value) or None,
            location=strip_prefix(clean(sheet.cell(row, 15).value), "Location:") or None,
            comment=strip_prefix(clean(sheet.cell(row, 16).value), "Comment:") or None,
            description=clean(sheet.cell(row, 18).value),
        )

    if not items:
        warnings.append(WarningMessage("wo.no_items", "No work-order items were found.", str(path)))

    return (
        WorkOrder(
            path=path,
            order_number=order_number,
            customer_name=customer_name,
            project_name=normalize_project_name(project_name),
            items=items,
        ),
        warnings,
    )


def expand_po_line(
    po_line: PurchaseOrderLine, work_item: WorkOrderItem | None, work_order: WorkOrder
) -> list[ExtractedJambLine]:
    del work_order
    attrs = parse_po_description(po_line.description)
    warnings: list[WarningMessage] = []
    if attrs["jamb_depth"] is None:
        warnings.append(WarningMessage("parse.no_depth", f"Could not parse jamb depth for {po_line.pella_xref}."))
    if attrs["finish_code"] is None:
        warnings.append(WarningMessage("parse.no_finish", f"Could not map finish for {po_line.pella_xref}."))

    units = unit_count(attrs, po_line.quantity, warnings, po_line)
    dimensions = dimensions_for_line(po_line, attrs, work_item, units, warnings)

    lines: list[ExtractedJambLine] = []
    for width, height, dim_warning in dimensions:
        line_warnings = list(warnings)
        if dim_warning:
            line_warnings.append(dim_warning)
        calculated_height = calculated_jamb_height(attrs["type"], height)
        footage = calculated_footage(attrs["type"], width, height)
        lines.append(
            ExtractedJambLine(
                po_line=po_line,
                work_item=work_item,
                type=attrs["type"],
                jamb_depth=attrs["jamb_depth"],
                jamb_width=round_reasonable(width),
                jamb_height=round_reasonable(height),
                quantity=1,
                piece_count=attrs["piece_count"],
                jamb_style=attrs["jamb_style"],
                lumber_type=attrs["lumber_type"],
                finish_text=attrs["finish_text"],
                finish_code=attrs["finish_code"],
                calculated_jamb_height=round_reasonable(calculated_height),
                calculated_footage=round_reasonable(footage),
                warnings=line_warnings,
            )
        )
    return lines


def parse_po_description(description: str) -> dict[str, object]:
    normalized = " ".join(description.replace("–", "-").split())
    lower = normalized.lower()
    piece_count = parse_piece_count(normalized)
    shape = "curved" if "curved" in lower else "arch" if "arch" in lower else None
    if shape == "curved":
        jamb_type = None
    elif "patio door" in lower or "3 pcs" in lower or shape == "arch":
        jamb_type = "3 Sided Patio Door"
    else:
        jamb_type = "4 Sided Window"
    jamb_style = "Drilled" if "drilled" in lower else "JE"
    lumber_type = next((material for material in LUMBER_TYPES if material.lower() in lower), None)
    finish_text = extract_finish_text(normalized)
    finish_code = map_finish(finish_text)

    depth_match = re.search(
        r"(?P<depth>\d+(?:\.\d+)?(?:[-\s]\d+/\d+)?|\d+/\d+)\s*\"?\s*\(Actual\)",
        normalized,
        flags=re.IGNORECASE,
    )
    jamb_depth = parse_number(depth_match.group("depth")) if depth_match else None

    return {
        "type": jamb_type,
        "jamb_depth": jamb_depth,
        "piece_count": piece_count,
        "jamb_style": jamb_style,
        "lumber_type": lumber_type,
        "finish_text": finish_text,
        "finish_code": finish_code,
        "shape": shape,
    }


def write_workbook(job: JobModel, output_path: str | Path) -> Path:
    from openpyxl import Workbook

    output_path = Path(output_path)
    workbook = Workbook()
    jamb_sheet = workbook.active
    jamb_sheet.title = "Jamb-BOM"
    lineal_sheet = workbook.create_sheet("Lineal-BOM")

    write_jamb_sheet(jamb_sheet, job)
    write_lineal_sheet(lineal_sheet, job)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def write_jamb_sheet(sheet, job: JobModel) -> None:
    sheet["G2"] = "Schedule"
    sheet["H2"] = job.schedule
    sheet["G3"] = "Jamb Thickness"
    sheet["H3"] = 0.6875
    sheet["G4"] = "Scribe Yes or No"
    sheet["H4"] = "No"
    sheet["G5"] = "Order Number(numbers only)"
    sheet["H5"] = job.order_number
    sheet["G6"] = "End Customer Name"
    sheet["H6"] = job.customer_name
    sheet["G7"] = "City"
    sheet["H7"] = job.ship_city
    sheet["G8"] = "State"
    sheet["H8"] = job.ship_state
    sheet["G9"] = "Pella PO Number"
    sheet["H9"] = job.pella_po_number
    sheet["G10"] = "Comment"
    sheet["G11"] = "Ship address"
    sheet["H11"] = job.ship_address
    sheet["G12"] = "Zip"
    sheet["H12"] = job.ship_zip
    sheet["G13"] = "Request Date"
    sheet["G14"] = "Date"
    sheet["H14"] = job.date_code

    for col, header in enumerate(INPUT_HEADERS, start=1):
        sheet.cell(18, col).value = header
    for col, header in enumerate(OUTPUT_HEADERS, start=20):
        sheet.cell(18, col).value = header

    row = 19
    for line in job.jamb_lines:
        write_input_line(sheet, row, line)
        if line.warnings:
            highlight_cells(sheet, row, 1, len(INPUT_HEADERS))
        row += 1

    out_row = 19
    debug = 1
    for line in job.jamb_lines:
        for loc in ("L", "R", "B", "T"):
            write_output_line(sheet, out_row, line, job, loc, debug)
            if line.warnings:
                highlight_cells(sheet, out_row, 20, 20 + len(OUTPUT_HEADERS) - 1)
            out_row += 1
        debug += 1

    format_sheet(sheet)


def write_input_line(sheet, row: int, line: ExtractedJambLine) -> None:
    values = [
        line.type,
        line.po_line.xref_suffix,
        line.po_line.pella_xref,
        None,
        None,
        line.jamb_depth,
        line.jamb_width,
        line.jamb_height,
        line.quantity,
        line.jamb_style,
        line.lumber_type,
        line.finish_code,
        line.calculated_jamb_height,
        line.calculated_footage,
    ]
    for col, value in enumerate(values, start=1):
        sheet.cell(row, col).value = value


def write_output_line(sheet, row: int, line: ExtractedJambLine, job: JobModel, loc: str, debug: int) -> None:
    length = length_for_loc(line, loc)
    part = "JC" if line.jamb_style == "Drilled" else "JE" if line.jamb_style else None
    finish = output_finish(line.lumber_type)
    values = [
        job.date_code,
        job.schedule,
        line.po_line.xref_suffix,
        debug,
        part,
        loc,
        line.jamb_depth,
        finish,
        line.quantity,
        length,
        job.order_number,
        line.finish_code,
        "Unassemble" if line.type == "3 Sided Patio Door" else None,
        debug,
        f"{line.lumber_type or ''} Jamb".strip(),
        job.customer_name,
        job.ship_city,
        job.ship_state,
        job.pella_po_number,
        job.order_number,
        None,
        job.ship_address,
        job.ship_zip,
        0,
        line.po_line.requested_date.isoformat() if line.po_line.requested_date else None,
        None,
        None,
    ]
    for col, value in enumerate(values, start=20):
        sheet.cell(row, col).value = value


def write_lineal_sheet(sheet, job: JobModel) -> None:
    del job
    for col, header in enumerate(LINEAL_HEADERS, start=1):
        sheet.cell(1, col).value = header
    format_sheet(sheet)


def highlight_cells(sheet, row: int, start_col: int, end_col: int) -> None:
    from openpyxl.styles import Border, PatternFill, Side

    fill = PatternFill("solid", fgColor="FFD966")
    border = Border(
        left=Side(style="thin", color="BF9000"),
        right=Side(style="thin", color="BF9000"),
        top=Side(style="thin", color="BF9000"),
        bottom=Side(style="thin", color="BF9000"),
    )
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = fill
        cell.border = border


def format_sheet(sheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row in {1, 18} or cell.column == 7:
                cell.font = bold
            if cell.row in {1, 18}:
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top")


def find_xref_index(row: list[str]) -> int | None:
    for index, value in enumerate(row):
        if XREF_RE.fullmatch(value.strip()):
            return index
    return None


def value_after(row: list[str], label: str) -> str | None:
    for index, value in enumerate(row):
        if value.strip().lower() == label.lower() and index + 1 < len(row):
            return row[index + 1].strip()
    return None


def parse_short_date(text: str) -> date | None:
    if not text:
        return None
    for fmt in ("%d-%b", "%d-%B", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            year = parsed.year if "%Y" in fmt else 2026
            return date(year, parsed.month, parsed.day)
        except ValueError:
            continue
    return None


def clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def strip_prefix(text: str, prefix: str) -> str:
    if text.lower().startswith(prefix.lower()):
        return text[len(prefix) :].strip()
    return text


def parse_int(value: object) -> int | None:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def normalize_project_name(name: str) -> str:
    if not name:
        return ""
    first = name.split(",", 1)[0].strip()
    return first.title() if first.isupper() else first


def parse_piece_count(description: str) -> int | None:
    match = re.search(r"(\d+)\s*pcs", description, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def extract_finish_text(description: str) -> str | None:
    star_match = re.findall(r"\*+\s*([^*]+?)\s*\*+", description)
    if star_match:
        for candidate in star_match:
            cleaned_text = candidate.strip().strip('"')
            if "longer" not in cleaned_text.lower() and cleaned_text:
                return cleaned_text

    paren_match = re.search(r"\((Unfinished|Primed)\s*\)", description, flags=re.IGNORECASE)
    if paren_match:
        return paren_match.group(1).strip()

    for key in FINISH_CODES:
        if key in description.lower():
            return key
    return None


def map_finish(finish_text: str | None) -> str | None:
    if not finish_text:
        return None
    cleaned_text = " ".join(finish_text.lower().replace("*", "").strip().split())
    if cleaned_text in FINISH_CODES:
        return FINISH_CODES[cleaned_text]
    for key, code in FINISH_CODES.items():
        if key in cleaned_text:
            return code
    return None


def unit_count(
    attrs: dict[str, object],
    po_quantity: int | None,
    warnings: list[WarningMessage],
    po_line: PurchaseOrderLine,
) -> int:
    piece_count = attrs["piece_count"]
    jamb_type = attrs["type"]
    if attrs.get("shape") == "curved" and isinstance(piece_count, int):
        return max(1, piece_count)
    sides = 3 if jamb_type == "3 Sided Patio Door" else 4
    if piece_count:
        units = piece_count / sides
        if units.is_integer():
            return max(1, int(units))
        warnings.append(
            WarningMessage(
                "parse.piece_count_mismatch",
                f"{po_line.pella_xref} has {piece_count} pcs, which is not divisible by {sides}.",
            )
        )
    return max(1, po_quantity or 1)


def dimensions_for_line(
    po_line: PurchaseOrderLine,
    attrs: dict[str, object],
    work_item: WorkOrderItem | None,
    units: int,
    warnings: list[WarningMessage],
) -> list[tuple[float | None, float | None, WarningMessage | None]]:
    del warnings
    if work_item is None:
        return [(None, None, WarningMessage("match.no_dimensions", f"No dimensions for {po_line.pella_xref}."))]

    description = work_item.description
    lower_po = po_line.description.lower()

    if attrs.get("shape") == "curved":
        return [
            (
                None,
                None,
                WarningMessage(
                    "parse.curved_unit",
                    f"{po_line.pella_xref} is a curved unit and dimensions must be reviewed.",
                ),
            )
            for _ in range(units)
        ]

    primary = primary_product_dimension(description)
    if primary and units == 1:
        return [(primary[0], primary[1], None)]

    components = component_dimensions(description)
    selected = select_components(lower_po, components, units)
    if not selected and primary:
        selected = [primary] * units
    elif not selected:
        rough = parse_dimension_pair(work_item.rough_opening or "")
        selected = [rough] * units if rough else []

    if not selected:
        return [(None, None, WarningMessage("parse.no_dimensions", f"Could not parse dimensions for {po_line.pella_xref}."))]

    results: list[tuple[float | None, float | None, WarningMessage | None]] = []
    for dimension in selected[:units]:
        warning = None
        width, height = dimension
        if "arch" in lower_po or "curved" in lower_po:
            warning = WarningMessage(
                "parse.complex_shape",
                f"{po_line.pella_xref} appears to be an arch/curved unit and should be reviewed.",
            )
        results.append((width, height, warning))
    while len(results) < units:
        results.append((None, None, WarningMessage("parse.missing_unit", f"Missing unit dimensions for {po_line.pella_xref}.")))
    return results


def primary_product_dimension(description: str) -> tuple[float, float] | None:
    head = description[:500]
    matches = list(re.finditer(r",\s*([^,]{0,40}?\d[^,]*?\s+[Xx]\s+[^,]*?\d)\s*,", head))
    if not matches:
        return None
    return parse_dimension_pair(matches[0].group(1))


def component_dimensions(description: str) -> list[tuple[str, float, float]]:
    matches: list[tuple[str, float, float]] = []
    pattern = re.compile(
        r"(?P<num>\d+):\s*(?P<label>.{0,140}?)Frame Size:\s*(?P<size>.{3,40}?)(?=\s+(?:Lifestyle|Architect|Impervia|Wood|Pella|General|Exterior|Interior))",
        flags=re.IGNORECASE | re.DOTALL,
    )
    for match in pattern.finditer(description):
        pair = parse_dimension_pair(match.group("size"))
        if pair:
            label = " ".join(match.group("label").split())
            matches.append((label, pair[0], pair[1]))
    return matches


def select_components(po_description_lower: str, components: list[tuple[str, float, float]], units: int) -> list[tuple[float, float]]:
    if not components:
        return []

    if "patio door" in po_description_lower:
        filtered = [(width, height) for label, width, height in components if "door" in label.lower()]
        return filtered[:units]

    if "arch" in po_description_lower or "curved" in po_description_lower:
        filtered = [(width, height) for label, width, height in components if "arch" in label.lower()]
        return filtered[:units]

    filtered = [
        (width, height)
        for label, width, height in components
        if "door" not in label.lower() and "arch" not in label.lower()
    ]
    return filtered[:units]


def calculated_jamb_height(jamb_type: str | None, height: float | None) -> float | None:
    if height is None:
        return None
    if jamb_type == "3 Sided Patio Door":
        return height - 0.875
    return height - (JAMB_THICKNESS * 2)


def calculated_footage(jamb_type: str | None, width: float | None, height: float | None) -> float | None:
    if width is None or height is None:
        return None
    if jamb_type == "3 Sided Patio Door":
        return (width + ((height + 2) * 2)) / 12
    return ((width * 2) + (height * 2)) / 12


def length_for_loc(line: ExtractedJambLine, loc: str) -> float | None:
    if loc in {"L", "R"}:
        return line.calculated_jamb_height
    if loc == "B" and line.type == "3 Sided Patio Door":
        return 0
    return line.jamb_width


def output_finish(lumber_type: str | None) -> str:
    if lumber_type == "Poplar":
        return "PAINT"
    if lumber_type == "Red Oak":
        return "SGO"
    if lumber_type == "Maple":
        return "SM"
    return "SPPN"


def parse_number(value: str | int | float | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace('"', "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        pass

    match = MIXED_FRACTION_RE.match(text)
    if not match:
        return None

    whole = int(match.group("whole") or 0)
    num = match.group("num")
    den = match.group("den")
    if num and den:
        denominator = int(den)
        if denominator == 0:
            return None
        return whole + int(num) / denominator
    return float(whole)


def parse_dimension_pair(text: str) -> tuple[float, float] | None:
    match = re.search(
        r"(?P<w>\d+(?:\.\d+)?(?:[-\s]\d+/\d+)?)\s*[Xx]\s*"
        r"(?P<h>\d+(?:\.\d+)?(?:[-\s]\d+/\d+)?)",
        text,
    )
    if not match:
        return None
    width = parse_number(match.group("w"))
    height = parse_number(match.group("h"))
    if width is None or height is None:
        return None
    return width, height


def round_reasonable(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 6)


def order_from_xref(xref: str) -> str:
    return xref.split("-", 1)[0] if xref else ""


def dedupe_warnings(warnings: list[WarningMessage]) -> list[WarningMessage]:
    unique = {}
    for warning in warnings:
        unique[(warning.code, warning.message)] = warning
    return list(unique.values())


if __name__ == "__main__":
    raise SystemExit(main())
