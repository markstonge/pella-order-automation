from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import WarningMessage, WorkOrder, WorkOrderItem


def parse_work_order_xlsx(path: str | Path) -> tuple[WorkOrder, list[WarningMessage]]:
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Version 1 only supports XLSX work orders: {path.name}")

    warnings: list[WarningMessage] = []
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    order_number = _clean(sheet["J1"].value)
    customer_name = _clean(sheet["G1"].value)
    project_name = _clean(sheet["H1"].value)

    if not order_number:
        warnings.append(WarningMessage("wo.no_order", "Could not find work-order number.", str(path)))

    items: dict[int, WorkOrderItem] = {}
    for row in range(2, sheet.max_row + 1):
        if _clean(sheet.cell(row, 1).value) != "Item No.":
            continue
        item_number = _parse_int(sheet.cell(row, 11).value)
        if item_number is None:
            continue
        items[item_number] = WorkOrderItem(
            source_row=row,
            item_number=item_number,
            quantity=_parse_int(sheet.cell(row, 13).value),
            rough_opening=_clean(sheet.cell(row, 8).value) or None,
            location=_strip_prefix(_clean(sheet.cell(row, 15).value), "Location:") or None,
            comment=_strip_prefix(_clean(sheet.cell(row, 16).value), "Comment:") or None,
            description=_clean(sheet.cell(row, 18).value),
        )

    if not items:
        warnings.append(WarningMessage("wo.no_items", "No work-order items were found.", str(path)))

    return (
        WorkOrder(
            path=path,
            order_number=order_number,
            customer_name=customer_name,
            project_name=_normalize_project_name(project_name),
            items=items,
        ),
        warnings,
    )


def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _strip_prefix(text: str, prefix: str) -> str:
    if text.lower().startswith(prefix.lower()):
        return text[len(prefix) :].strip()
    return text


def _parse_int(value: object) -> int | None:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _normalize_project_name(name: str) -> str:
    if not name:
        return ""
    first = name.split(",", 1)[0].strip()
    return first.title() if first.isupper() else first
