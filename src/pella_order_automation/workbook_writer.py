from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ExtractedJambLine, JobModel


INPUT_HEADERS = [
    "Type",
    "PO Line",
    "Pella Xref",
    "Wall Depth(JambDepth)",
    "Window Depth",
    "Jamb Depth",
    "Jamb Width",
    "Jamb height",
    "Quantity",
    "Jamb Style",
    "Lumber Type",
    "Finish",
    "Calculated Jamb height",
    "Calculated Footage",
]

OUTPUT_HEADERS = [
    "Date",
    "Schedule",
    "Batch(Box =)",
    "Bin",
    "Part",
    "Loc",
    "W",
    "Fin",
    "Qty",
    "L",
    "Info",
    "Order NumLer",
    "Assembled",
    "Debug",
    "Part No",
    "Customer",
    "Shp Addr City",
    "Shp Addr State",
    "PONumber",
    "Customer Ref",
    "Comment",
    "Shp Addr Address1",
    "Shp Addr Zip Code",
    "orderline",
    "Req Date",
    "SUMP",
    "Shim SUMP",
]

LINEAL_HEADERS = [
    "Date",
    "Schedule",
    "Batch(Box #)",
    "Bin",
    "Part",
    "Loc",
    "W",
    "Fin",
    "Qty",
    "L",
    "Info",
    "Part No",
    "Customer",
    "Shp Addr City",
    "Shp Addr State",
    "PONumber",
    "Customer Ref",
    "Comment",
    "Shp Addr Add1",
    "Shp Addr Zip",
    "Order Number",
    "Line Item",
    "Req Date",
    "SUMP",
]

REVIEW_FILL = PatternFill("solid", fgColor="FFD966")
REVIEW_BORDER = Border(
    left=Side(style="thin", color="BF9000"),
    right=Side(style="thin", color="BF9000"),
    top=Side(style="thin", color="BF9000"),
    bottom=Side(style="thin", color="BF9000"),
)


def write_workbook(job: JobModel, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    workbook = Workbook()
    jamb_sheet = workbook.active
    jamb_sheet.title = "Jamb-BOM"
    lineal_sheet = workbook.create_sheet("Lineal-BOM")

    _write_jamb_sheet(jamb_sheet, job)
    _write_lineal_sheet(lineal_sheet, job)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _write_jamb_sheet(sheet, job: JobModel) -> None:
    sheet["G2"] = "Schedule"
    sheet["H2"] = job.schedule
    sheet["G3"] = "Jamb Thickness"
    sheet["H3"] = 0.6875
    sheet["G4"] = "Scribe Yes or No"
    sheet["H4"] = "No"
    sheet["G5"] = "Order Number(numbers only)"
    sheet["H5"] = job.order_number
    sheet["G6"] = "End Customer Name"
    sheet["H6"] = job.customer_name
    sheet["G7"] = "City"
    sheet["H7"] = job.ship_city
    sheet["G8"] = "State"
    sheet["H8"] = job.ship_state
    sheet["G9"] = "Pella PO Number"
    sheet["H9"] = job.pella_po_number
    sheet["G10"] = "Comment"
    sheet["G11"] = "Ship address"
    sheet["H11"] = job.ship_address
    sheet["G12"] = "Zip"
    sheet["H12"] = job.ship_zip
    sheet["G13"] = "Request Date"
    sheet["G14"] = "Date"
    sheet["H14"] = job.date_code

    for col, header in enumerate(INPUT_HEADERS, start=1):
        sheet.cell(18, col).value = header
    for col, header in enumerate(OUTPUT_HEADERS, start=20):
        sheet.cell(18, col).value = header

    row = 19
    for line in job.jamb_lines:
        _write_input_line(sheet, row, line)
        if line.warnings:
            _highlight_cells(sheet, row, 1, len(INPUT_HEADERS))
        row += 1

    out_row = 19
    debug = 1
    for line in job.jamb_lines:
        for loc in ("L", "R", "B", "T"):
            _write_output_line(sheet, out_row, line, job, loc, debug)
            if line.warnings:
                _highlight_cells(sheet, out_row, 20, 20 + len(OUTPUT_HEADERS) - 1)
            out_row += 1
        debug += 1

    _format_sheet(sheet)


def _write_input_line(sheet, row: int, line: ExtractedJambLine) -> None:
    values = [
        line.type,
        line.po_line.xref_suffix,
        line.po_line.pella_xref,
        None,
        None,
        line.jamb_depth,
        line.jamb_width,
        line.jamb_height,
        line.quantity,
        line.jamb_style,
        line.lumber_type,
        line.finish_code,
        line.calculated_jamb_height,
        line.calculated_footage,
    ]
    for col, value in enumerate(values, start=1):
        sheet.cell(row, col).value = value


def _write_output_line(sheet, row: int, line: ExtractedJambLine, job: JobModel, loc: str, debug: int) -> None:
    length = _length_for_loc(line, loc)
    part = "JC" if line.jamb_style == "Drilled" else "JE" if line.jamb_style else None
    finish = _output_finish(line.lumber_type)
    values = [
        job.date_code,
        job.schedule,
        line.po_line.xref_suffix,
        debug,
        part,
        loc,
        line.jamb_depth,
        finish,
        line.quantity,
        length,
        job.order_number,
        line.finish_code,
        "Unassemble" if line.type == "3 Sided Patio Door" else None,
        debug,
        f"{line.lumber_type or ''} Jamb".strip(),
        job.customer_name,
        job.ship_city,
        job.ship_state,
        job.pella_po_number,
        job.order_number,
        None,
        job.ship_address,
        job.ship_zip,
        0,
        line.po_line.requested_date.isoformat() if line.po_line.requested_date else None,
        None,
        None,
    ]
    for col, value in enumerate(values, start=20):
        sheet.cell(row, col).value = value


def _length_for_loc(line: ExtractedJambLine, loc: str) -> float | None:
    if loc in {"L", "R"}:
        return line.calculated_jamb_height
    if loc == "B" and line.type == "3 Sided Patio Door":
        return 0
    return line.jamb_width


def _output_finish(lumber_type: str | None) -> str:
    if lumber_type == "Poplar":
        return "PAINT"
    if lumber_type == "Red Oak":
        return "SGO"
    if lumber_type == "Maple":
        return "SM"
    return "SPPN"


def _write_lineal_sheet(sheet, job: JobModel) -> None:
    for col, header in enumerate(LINEAL_HEADERS, start=1):
        sheet.cell(1, col).value = header
    _format_sheet(sheet)


def _highlight_cells(sheet, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = REVIEW_FILL
        cell.border = REVIEW_BORDER


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row in {1, 18} or cell.column == 7:
                cell.font = bold
            if cell.row in {1, 18}:
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top")

    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(str(cell.value)) + 2), 36)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = max(10, width)
    sheet.freeze_panes = "A19" if sheet.title == "Jamb-BOM" else "A2"
